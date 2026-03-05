from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, g
from sqlalchemy.orm import joinedload
from sqlalchemy import or_, func, and_
from .db import get_db
from .models import Ride, User, Booking, CompletedRide, Rating, Report, AdminUser, RideMessage, UserBlock
from .auth import load_current_user, login_required, admin_required
from flask import abort
from pathlib import Path
import openpyxl
import re

# I used ChatGPT to help read the xlsx file for the venues and artists. I prompted it with "how would i be able to connect an xlsx file to python'

DATA_DIR = Path(__file__).resolve().parent / "static" / "data"
def load_artists():
    xlsx = DATA_DIR / "artists.xlsx"
    rows = []
    if not xlsx.exists():
        return rows

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        i_name = header.index("name")
        i_genre = header.index("genre")
    except ValueError:
        return rows

    for r in ws.iter_rows(min_row=2, values_only=True):
        name = (r[i_name] or "").strip() if r[i_name] else ""
        genre = (r[i_genre] or "").strip() if r[i_genre] else ""
        if name:
            rows.append({"name": name, "genre": genre})
    return rows

def load_venues():
    rows = []
    candidates = [DATA_DIR / "venues_ie.csv", DATA_DIR / "venues_ie.xlsx"]
    file = next((p for p in candidates if p.exists()), None)
    rows: list[dict] = []
    if not file:
        return rows

    if file.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        header = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [str(h).strip().lower() for h in header]
        try:
            i_name = header.index("name")
            i_county = header.index("county")
        except ValueError:
            return rows
        for r in ws.iter_rows(min_row=2, values_only=True):
            name = (r[i_name] or "").strip() if r[i_name] else ""
            county = (r[i_county] or "").strip() if r[i_county] else ""
            if name and county:
                rows.append({"name": name, "county": county})
    return rows

def load_counties():
    try:
        return sorted({v["county"] for v in load_venues()})
    except Exception:
        return []

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def is_valid_email(email: str) -> bool:
    return bool(EMAIL_RE.match(email or ""))

def password_error(pw: str):
    pw = pw or ""
    if len(pw) < 8:
        return "Password must be at least 8 characters."
    if not any(c.isupper() for c in pw):
        return "Password must include at least one capital letter."
    if not any(c.isdigit() for c in pw):
        return "Password must include at least one number."
    return None

def has_block_between(db, a_id: int, b_id: int) -> bool:
    """True if a blocks b OR b blocks a."""
    return db.query(UserBlock).filter(
        or_(
            and_(UserBlock.blocker_id == a_id, UserBlock.blocked_id == b_id),
            and_(UserBlock.blocker_id == b_id, UserBlock.blocked_id == a_id),
        )
    ).first() is not None
app = Flask(__name__)
app.secret_key = "122341116"

@app.before_request
def _load_user():
    load_current_user()
    g.is_admin = False
    if g.user:
        [db] = list(get_db())
        try:
            g.is_admin = db.query(AdminUser).get(g.user.id) is not None
        finally:
            db.close()

@app.context_processor
def inject_user():
    return {"current_user": getattr(g, "user", None)}

#Homepage
@app.get("/")
def home():
    [db] = list(get_db())
    try:
        f_artist = (request.args.get("artist") or "").strip()
        f_county = (request.args.get("county") or "").strip()
        f_tags   = [t.strip().lower() for t in request.args.getlist("tags")]
        f_avail  = request.args.get("available") == "1"
        f_from   = (request.args.get("from") or "").strip()
        f_to     = (request.args.get("to") or "").strip()
        dt_from = datetime.strptime(f_from, "%Y-%m-%d") if f_from else None
        dt_to   = datetime.strptime(f_to, "%Y-%m-%d") if f_to else None
        artists  = [a["name"] for a in load_artists()] if "load_artists" in globals() else []
        counties = load_counties()
        rides = (
            db.query(Ride)
              .options(joinedload(Ride.driver), joinedload(Ride.bookings))
              .order_by(Ride.depart_at.asc())
              .all()
        )

        blocked_user_ids = set()
        if g.user:
            blocks = db.query(UserBlock).filter(
                (UserBlock.blocker_id == g.user.id) | (UserBlock.blocked_id == g.user.id)
            ).all()
            blocked_user_ids = {b.blocked_id for b in blocks if b.blocker_id == g.user.id} | \
                               {b.blocker_id for b in blocks if b.blocked_id == g.user.id}

        finished_ids = {rid for (rid,) in db.query(CompletedRide.ride_id).all()}

        joined_ids = set()
        if g.user:
            joined_ids = {
                b.ride_id
                for b in db.query(Booking).filter(Booking.rider_id == g.user.id).all()
            }

        def has_available_seat(r): return len(r.bookings) < r.seats_total
        def county_in_destination(r, county):
            return county.lower() in (r.destination or "").lower()

# I used ChatGPT to help create the filters for the homepage. I copied over the web.py and ride_detail.html and prompted it with "how would i be able to create filters for this code"

        filtered = []
        for r in rides:
            # I used ChatGPT to help hide the rides from the home page. I copied over web.py and prompted it with "how can i make so the rides disapear from the homepage after they are finished
            if r.id in finished_ids:
                continue

            if f_artist and f_artist.lower() not in (r.artist or "").lower():
                continue
            if f_county and not county_in_destination(r, f_county):
                continue
            if dt_from and r.depart_at < dt_from:
                continue
            if dt_to and r.depart_at > dt_to:
                continue
            if f_tags:
                ride_tags = [t.lower() for t in r.tags]
                if not any(t in ride_tags for t in f_tags):
                    continue
            if f_avail and not has_available_seat(r):
                continue
            filtered.append(r)
            if blocked_user_ids and r.driver_id in blocked_user_ids:
                continue

        return render_template(
            "rides_list.html",
            rides=filtered,
            joined_ids=joined_ids,
            finished_ids=finished_ids,
            artists=artists,
            counties=counties,
            sel=dict(artist=f_artist, county=f_county, tags=f_tags,
                     available=f_avail, date_from=f_from, date_to=f_to),
        )
    finally:
        db.close()


# Signup
@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        return render_template("signup.html")

    name = request.form.get("name","").strip()
    email = request.form.get("email","").strip().lower()
    password = request.form.get("password","")

    if not all([name, email, password]):
        flash("All fields are required.")
        return redirect(url_for("signup"))

    if not is_valid_email(email):
        flash("Please enter a valid email address.")
        return redirect(url_for("signup"))

    err = password_error(password)
    if err:
        flash(err)
        return redirect(url_for("signup"))

    [db] = list(get_db())
    try:
        if db.query(User).filter(User.email == email).first():
            flash("Email already registered. Try logging in.")
            return redirect(url_for("login"))

        user = User(name=name, email=email, password=password)
        db.add(user)
        db.commit()

        if db.query(AdminUser).count() == 0:
            db.add(AdminUser(user_id=user.id))
            db.commit()

        db.refresh(user)
        session["user_id"] = user.id
        flash("Welcome! Account created.")
        return redirect(url_for("home"))
    finally:
        db.close()

# Login
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    email = request.form.get("email","").strip().lower()
    password = request.form.get("password","")
    [db] = list(get_db())
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user or user.password != password:
            flash("Invalid email or password.")
            return redirect(url_for("login"))
        session["user_id"] = user.id
        flash("Logged in.")
        if db.query(AdminUser).get(user.id):
            return redirect(url_for("admin_home"))
        return redirect(url_for("home"))
    finally:
        db.close()

# Logout
@app.get("/logout")
def logout():
    session.pop("user_id", None)
    flash("Logged out.")
    return redirect(url_for("home"))

# Creating Rides
@app.route("/rides/new", methods=["GET", "POST"])
@login_required
def rides_new():
    if request.method == "GET":
        venues = load_venues()
        artists = load_artists()
        return render_template("ride_new.html", venues=venues, artists=artists)

    destination = request.form.get("destination","").strip()
    origin = request.form.get("origin", "").strip()
    depart_at_str = request.form.get("depart_at","").strip()
    seats_total = request.form.get("seats_total","3").strip()
    artist = request.form.get("artist", "").strip()
    tags = request.form.getlist("tags")
    origin_lat = request.form.get("origin_lat")
    origin_lng = request.form.get("origin_lng")

    lat = float(origin_lat) if origin_lat else None
    lng = float(origin_lng) if origin_lng else None

    if lat is None or lng is None:
        flash("Please pick a starting location on the map.")
        return redirect(url_for("rides_new"))

    if not artist:
        flash("Artist is required.")
        return redirect(url_for("rides_new"))

    if not all([origin, destination, depart_at_str]):
        flash("All fields are required.")
        return redirect(url_for("rides_new"))

    try:
        depart_at = datetime.strptime(depart_at_str, "%Y-%m-%dT%H:%M")
        seats_total = int(seats_total)
        if seats_total < 1:
            raise ValueError
    except Exception:
        flash("Invalid date/time or seats.")
        return redirect(url_for("rides_new"))

    [db] = list(get_db())
    try:
        ride = Ride(
            driver_id=g.user.id,
            origin=origin,
            destination=destination,
            depart_at=depart_at,
            seats_total=seats_total,
            artist=artist,
            tags=tags,
            origin_lat=lat,
            origin_lng=lng,
        )
        db.add(ride)
        db.commit()
        flash("Ride created")
        return redirect(url_for("home"))
    finally:
        db.close()



# Joining Rides
@app.post("/rides/<int:ride_id>/join")
@login_required
def join_ride(ride_id):
    [db] = list(get_db())
    try:
        ride = (
            db.query(Ride)
              .options(joinedload(Ride.bookings), joinedload(Ride.driver))
              .get(ride_id)
        )
        if not ride:
            flash("Ride not found.")
            return redirect(url_for("home"))

        if db.query(CompletedRide).get(ride_id) is not None:
            flash("This ride has already been finished.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        if ride.driver_id == g.user.id:
            flash("You cant join your own ride.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        already = db.query(Booking).filter(
            Booking.ride_id == ride.id, Booking.rider_id == g.user.id
        ).first()
        if already:
            flash("You already joined this ride.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        seats_taken = len(ride.bookings)
        if seats_taken >= ride.seats_total:
            flash("Sorry this ride is full.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

# With the help of ChatGPT I made it so users can't join blocked rides. I copied web.py over and prompted it with 'how would i be able to make it so users cant join blocked users rides'
        participant_ids = {ride.driver_id} | {b.rider_id for b in ride.bookings}
        for pid in participant_ids:
            if has_block_between(db, g.user.id, pid):
                flash("You can’t join this ride because one of the participants has blocked the other.")
                return redirect(url_for("ride_detail", ride_id=ride_id))

        db.add(Booking(ride_id=ride.id, rider_id=g.user.id))
        db.commit()
        flash("You’ve joined the ride")
        return redirect(url_for("ride_detail", ride_id=ride_id))
    finally:
        db.close()

# Canceling Rides
@app.post("/rides/<int:ride_id>/cancel")
@login_required
def cancel_join(ride_id):
    [db] = list(get_db())
    try:
        b = db.query(Booking).filter(
            Booking.ride_id == ride_id, Booking.rider_id == g.user.id
        ).first()
        if not b:
            flash("You’re not booked on that ride.")
            return redirect(url_for("home"))
        db.delete(b)
        db.commit()
        flash("Booking cancelled.")
        return redirect(url_for("home"))
    finally:
        db.close()

# Bookings
@app.get("/my/bookings")
@login_required
def my_bookings():
    [db] = list(get_db())
    try:
        bookings = (
            db.query(Booking)
              .options(joinedload(Booking.ride).joinedload(Ride.driver))
              .filter(Booking.rider_id == g.user.id)
              .order_by(Ride.depart_at.asc())
              .all()
        )
        return render_template("my_bookings.html", bookings=bookings)
    finally:
        db.close()

# Ride Details
@app.get("/rides/<int:ride_id>")
def ride_detail(ride_id):
    [db] = list(get_db())
    try:
        ride = (
            db.query(Ride)
              .options(
                  joinedload(Ride.driver),
                  joinedload(Ride.bookings).joinedload(Booking.rider),
              )
              .get(ride_id)
        )
        if not ride:
            flash("Ride not found.")
            return redirect(url_for("home"))

        is_finished = db.query(CompletedRide).get(ride_id) is not None
        is_driver = g.user and g.user.id == ride.driver_id

        # With the help of ChatGPT I made it users can't see messages from blocked users. I prompted chatgpt with 'how would i be able to block user messages from blocked users
        messages = []
        is_participant = False
        my_blocked_ids = set()
        blocked_me_ids = set()

        if g.user:
            participant_ids = {ride.driver_id} | {b.rider_id for b in ride.bookings}
            is_participant = g.user.id in participant_ids

            if is_participant:
                blocks = db.query(UserBlock).filter(
                    or_(UserBlock.blocker_id == g.user.id, UserBlock.blocked_id == g.user.id)
                ).all()

                my_blocked_ids = {b.blocked_id for b in blocks if b.blocker_id == g.user.id}
                blocked_me_ids = {b.blocker_id for b in blocks if b.blocked_id == g.user.id}
                hidden_sender_ids = my_blocked_ids | blocked_me_ids

                messages = (
                    db.query(RideMessage)
                      .options(joinedload(RideMessage.sender))
                      .filter(RideMessage.ride_id == ride_id)
                      .order_by(RideMessage.created_at.asc())
                      .all()
                )
                messages = [m for m in messages if m.sender_id not in hidden_sender_ids]

        return render_template(
            "ride_detail.html",
            ride=ride,
            is_driver=is_driver,
            is_finished=is_finished,
            is_participant=is_participant,
            messages=messages,
            my_blocked_ids=my_blocked_ids,
            blocked_me_ids=blocked_me_ids,
        )
    finally:
        db.close()


#Kicking Users
#This Kicking Users code is from chatgpt, I uploaded web.py, base.html and ride_detail.html and asked "how i would be able to add a kick users button"
@app.post("/rides/<int:ride_id>/kick/<int:booking_id>")
@login_required
def kick_passenger(ride_id, booking_id):
    [db] = list(get_db())
    try:
        ride = db.query(Ride).get(ride_id)
        if not ride:
            flash("Ride not found.")
            return redirect(url_for("home"))

        if ride.driver_id != g.user.id:
            abort(403)

        booking = (
            db.query(Booking)
              .options(joinedload(Booking.rider))
              .filter(Booking.id == booking_id, Booking.ride_id == ride_id)
              .first()
        )
        if not booking:
            flash("Booking not found.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        if booking.rider_id == ride.driver_id:
            flash("You can’t kick yourself.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        db.delete(booking)
        db.commit()
        flash(f"Removed {booking.rider.name} from the ride.")
        return redirect(url_for("ride_detail", ride_id=ride_id))
    finally:
        db.close()

#Profile
@app.get("/profile")
@login_required
def profile():
    [db] = list(get_db())
    try:
        # I used ChatGPT help add the past rides to the profile pages. I copied over web.py and prompted it with "How would i be able to move the finished rides to the profile pages"
        past_rides = (
            db.query(Ride)
              .join(CompletedRide, CompletedRide.ride_id == Ride.id)
              .outerjoin(Booking, Booking.ride_id == Ride.id)
              .filter(or_(Ride.driver_id == g.user.id, Booking.rider_id == g.user.id))
              .distinct()
              .order_by(CompletedRide.completed_at.desc())
              .all()
        )

        ratings = (
            db.query(Rating)
              .filter(Rating.ratee_id == g.user.id)
              .order_by(Rating.created_at.desc())
              .all()
        )

        avg = db.query(func.avg(Rating.stars)).filter(Rating.ratee_id == g.user.id).scalar()
        avg = float(avg) if avg is not None else None

        return render_template("profile.html", user=g.user, past_rides=past_rides, ratings=ratings, avg_rating=avg)
    finally:
        db.close()

#Profile Update
@app.post("/profile")
@login_required
def profile_update():
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip().lower()
    current_pw = request.form.get("current_password", "")
    new_pw = request.form.get("new_password", "")
    confirm_pw = request.form.get("confirm_password", "")

    if not name or not email:
        flash("Name and email are required.")
        return redirect(url_for("profile"))

    if not is_valid_email(email):
        flash("Please enter a valid email address.")
        return redirect(url_for("profile"))

    # I used ChatGPT to help fix a problem with creating a new password. I copied over web.py and profile.html and prompted it with "the new password doesnt enter the database when changed"
    wants_password_change = bool(new_pw or confirm_pw)

    [db] = list(get_db())
    try:
        existing = db.query(User).filter(User.email == email).first()
        if existing and existing.id != g.user.id:
            flash("That email is already in use.")
            return redirect(url_for("profile"))

        u = db.query(User).get(g.user.id)
        u.name = name
        u.email = email

        password_changed = False

        if wants_password_change:
            if not current_pw:
                flash("Enter your current password to change it.")
                return redirect(url_for("profile"))
            if current_pw != u.password:
                flash("Current password is incorrect.")
                return redirect(url_for("profile"))

            err = password_error(new_pw)
            if err:
                flash(err)
                return redirect(url_for("profile"))

            if new_pw != confirm_pw:
                flash("New password and confirmation don’t match.")
                return redirect(url_for("profile"))

            u.password = new_pw
            password_changed = True

        db.commit()

        flash("New password saved." if password_changed else "Profile updated.")
        return redirect(url_for("profile"))
    finally:
        db.close()
# Finish Ride
@app.post("/rides/<int:ride_id>/finish")
@login_required
def finish_ride(ride_id):
    [db] = list(get_db())
    try:
        ride = db.query(Ride).get(ride_id)
        if not ride:
            flash("Ride not found.")
            return redirect(url_for("home"))

        if ride.driver_id != g.user.id:
            abort(403)

        already = db.query(CompletedRide).get(ride_id)
        if already:
            flash("Ride is already finished.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        db.add(CompletedRide(ride_id=ride_id))
        db.commit()
        flash("Ride finished. It’s now in Past Rides.")
        return redirect(url_for("ride_detail", ride_id=ride_id))
    finally:
        db.close()
#Ride / Report Page
@app.get("/rides/<int:ride_id>/review")
@login_required
def review_ride(ride_id):
    [db] = list(get_db())
    try:
        ride = (
            db.query(Ride)
              .options(joinedload(Ride.driver), joinedload(Ride.bookings).joinedload(Booking.rider))
              .get(ride_id)
        )
        if not ride:
            flash("Ride not found.")
            return redirect(url_for("profile"))

        if not db.query(CompletedRide).get(ride_id):
            flash("That ride isn’t finished yet.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        participant_ids = {ride.driver_id} | {b.rider_id for b in ride.bookings}
        if g.user.id not in participant_ids:
            abort(403)

        participants = [ride.driver] + [b.rider for b in ride.bookings]

        existing = (
            db.query(Rating)
              .filter(Rating.ride_id == ride_id, Rating.rater_id == g.user.id)
              .all()
        )
        existing_map = {(r.ratee_id): r for r in existing}

        return render_template("ride_review.html", ride=ride, participants=participants, existing_map=existing_map)
    finally:
        db.close()
# Rate Submission
@app.post("/rides/<int:ride_id>/rate/<int:user_id>")
@login_required
def rate_user(ride_id, user_id):
    stars = int(request.form.get("stars", "0") or 0)
    comment = (request.form.get("comment", "") or "").strip()

    if stars < 1 or stars > 5:
        flash("Stars must be between 1 and 5.")
        return redirect(url_for("review_ride", ride_id=ride_id))

    [db] = list(get_db())
    try:
        ride = db.query(Ride).options(joinedload(Ride.bookings)).get(ride_id)
        if not ride or not db.query(CompletedRide).get(ride_id):
            flash("Ride not found or not finished.")
            return redirect(url_for("profile"))

        participant_ids = {ride.driver_id} | {b.rider_id for b in ride.bookings}
        if g.user.id not in participant_ids or user_id not in participant_ids or user_id == g.user.id:
            abort(403)

        r = db.query(Rating).filter_by(ride_id=ride_id, rater_id=g.user.id, ratee_id=user_id).first()
        if r:
            r.stars = stars
            r.comment = comment
            flash("Rating updated.")
        else:
            db.add(Rating(ride_id=ride_id, rater_id=g.user.id, ratee_id=user_id, stars=stars, comment=comment))
            flash("Rating saved.")

        db.commit()
        return redirect(url_for("review_ride", ride_id=ride_id))
    finally:
        db.close()
# Report Submission
@app.post("/rides/<int:ride_id>/report/<int:user_id>")
@login_required
def report_user(ride_id, user_id):
    reason = (request.form.get("reason", "") or "").strip()
    comment = (request.form.get("comment", "") or "").strip()

    if not reason:
        flash("Please add a reason for the report.")
        return redirect(url_for("review_ride", ride_id=ride_id))

    [db] = list(get_db())
    try:
        ride = db.query(Ride).options(joinedload(Ride.bookings)).get(ride_id)
        if not ride or not db.query(CompletedRide).get(ride_id):
            flash("Ride not found or not finished.")
            return redirect(url_for("profile"))

        participant_ids = {ride.driver_id} | {b.rider_id for b in ride.bookings}
        if g.user.id not in participant_ids or user_id not in participant_ids or user_id == g.user.id:
            abort(403)

        rep = db.query(Report).filter_by(ride_id=ride_id, reporter_id=g.user.id, reported_user_id=user_id).first()
        if rep:
            rep.reason = reason
            rep.comment = comment
            flash("Report updated.")
        else:
            db.add(Report(ride_id=ride_id, reporter_id=g.user.id, reported_user_id=user_id, reason=reason, comment=comment))
            flash("Report submitted.")

        db.commit()
        return redirect(url_for("review_ride", ride_id=ride_id))
    finally:
        db.close()

#Public Profile Page
@app.get("/users/<int:user_id>")
def profile_public(user_id):
    [db] = list(get_db())
    try:
        u = db.query(User).get(user_id)
        if not u:
            flash("User not found.")
            return redirect(url_for("home"))

        ratings = (
            db.query(Rating)
              .filter(Rating.ratee_id == user_id)
              .order_by(Rating.created_at.desc())
              .all()
        )
        avg = db.query(func.avg(Rating.stars)).filter(Rating.ratee_id == user_id).scalar()
        avg = float(avg) if avg is not None else None

        return render_template("profile_public.html", u=u, ratings=ratings, avg_rating=avg)
    finally:
        db.close()

#Admin home page
@app.get("/admin")
@login_required
@admin_required
def admin_home():
    [db] = list(get_db())
    try:
        user_count = db.query(func.count(User.id)).scalar() or 0
        report_count = db.query(func.count(Report.id)).scalar() or 0
        total_bookings = db.query(func.count(Booking.id)).scalar() or 0
        total_finished_rides = db.query(func.count(CompletedRide.ride_id)).scalar() or 0
        total_rides = db.query(func.count(Ride.id)).scalar() or 0

        return render_template(
            "admin.html",
            user_count=user_count,
            report_count=report_count,
            total_bookings=total_bookings,
            total_finished_rides=total_finished_rides,
            total_rides=total_rides,
        )
    finally:
        db.close()

# Admin report page
@app.get("/admin/reports")
@login_required
@admin_required
def admin_reports():
    [db] = list(get_db())
    try:
        reports = db.query(Report).order_by(Report.created_at.desc()).all()

        user_ids = set()
        ride_ids = set()
        for r in reports:
            user_ids.update([r.reporter_id, r.reported_user_id])
            ride_ids.add(r.ride_id)

        users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
        rides = db.query(Ride).filter(Ride.id.in_(ride_ids)).all() if ride_ids else []

        user_map = {u.id: u for u in users}
        ride_map = {ri.id: ri for ri in rides}

        return render_template(
            "admin_reports.html",
            reports=reports,
            user_map=user_map,
            ride_map=ride_map
        )
    finally:
        db.close()

# Admin delete page

@app.post("/admin/users/<int:user_id>/delete")
@login_required
@admin_required
def admin_delete_user(user_id):
    [db] = list(get_db())
    try:
        u = db.query(User).get(user_id)
        if not u:
            flash("User not found.")
            return redirect(url_for("admin_home"))

        if u.id == g.user.id:
            flash("You cannot delete your own admin account while logged in.")
            return redirect(url_for("admin_home"))

        db.query(AdminUser).filter(AdminUser.user_id == user_id).delete(synchronize_session=False)

        db.query(Report).filter(
            (Report.reporter_id == user_id) | (Report.reported_user_id == user_id)
        ).delete(synchronize_session=False)

        db.query(Rating).filter(
            (Rating.rater_id == user_id) | (Rating.ratee_id == user_id)
        ).delete(synchronize_session=False)

        db.query(Booking).filter(Booking.rider_id == user_id).delete(synchronize_session=False)

        ride_ids = [rid for (rid,) in db.query(Ride.id).filter(Ride.driver_id == user_id).all()]
        if ride_ids:
            db.query(Booking).filter(Booking.ride_id.in_(ride_ids)).delete(synchronize_session=False)
            db.query(CompletedRide).filter(CompletedRide.ride_id.in_(ride_ids)).delete(synchronize_session=False)
            db.query(Rating).filter(Rating.ride_id.in_(ride_ids)).delete(synchronize_session=False)
            db.query(Report).filter(Report.ride_id.in_(ride_ids)).delete(synchronize_session=False)
            db.query(Ride).filter(Ride.id.in_(ride_ids)).delete(synchronize_session=False)

        db.delete(u)
        db.commit()

        flash("User deleted from website and database.")
        return redirect(url_for("admin_home"))
    finally:
        db.close()

# Blocking Users
@app.post("/users/<int:user_id>/block")
@login_required
def block_user(user_id):
    if user_id == g.user.id:
        flash("You can’t block yourself.")
        return redirect(request.referrer or url_for("profile"))

    [db] = list(get_db())
    try:
        u = db.query(User).get(user_id)
        if not u:
            flash("User not found.")
            return redirect(url_for("home"))

        existing = db.query(UserBlock).filter_by(blocker_id=g.user.id, blocked_id=user_id).first()
        if existing:
            flash("User already blocked.")
            return redirect(request.referrer or url_for("profile_public", user_id=user_id))

        db.add(UserBlock(blocker_id=g.user.id, blocked_id=user_id))


        shared_as_driver = (
            db.query(Booking)
              .join(Ride, Ride.id == Booking.ride_id)
              .filter(Ride.driver_id == g.user.id, Booking.rider_id == user_id)
              .all()
        )

        shared_as_passenger = (
            db.query(Booking)
              .join(Ride, Ride.id == Booking.ride_id)
              .filter(Ride.driver_id == user_id, Booking.rider_id == g.user.id)
              .all()
        )

        for bk in shared_as_driver + shared_as_passenger:
            db.delete(bk)

        db.commit()
        flash(f"Blocked {u.name}.")
        return redirect(request.referrer or url_for("profile_public", user_id=user_id))

    finally:
        db.close()

# Unblocking Users
@app.post("/users/<int:user_id>/unblock")
@login_required
def unblock_user(user_id):
    [db] = list(get_db())
    try:
        b = db.query(UserBlock).filter_by(blocker_id=g.user.id, blocked_id=user_id).first()
        if not b:
            flash("That user isn’t blocked.")
        else:
            db.delete(b)
            db.commit()
            flash("User unblocked.")
        return redirect(request.referrer or url_for("profile_public", user_id=user_id))
    finally:
        db.close()

#Messaging
@app.post("/rides/<int:ride_id>/messages")
@login_required
def post_ride_message(ride_id):
    body = (request.form.get("body") or "").strip()
    if not body:
        flash("Message can’t be empty.")
        return redirect(url_for("ride_detail", ride_id=ride_id))
    if len(body) > 1000:
        flash("Message is too long (max 1000 chars).")
        return redirect(url_for("ride_detail", ride_id=ride_id))

    [db] = list(get_db())
    try:
        ride = db.query(Ride).options(joinedload(Ride.bookings)).get(ride_id)
        if not ride:
            flash("Ride not found.")
            return redirect(url_for("home"))

        if db.query(CompletedRide).get(ride_id) is not None:
            flash("This ride is finished. Chat is read-only.")
            return redirect(url_for("ride_detail", ride_id=ride_id))

        participant_ids = {ride.driver_id} | {b.rider_id for b in ride.bookings}
        if g.user.id not in participant_ids:
            abort(403)

        db.add(RideMessage(ride_id=ride_id, sender_id=g.user.id, body=body))
        db.commit()
        return redirect(url_for("ride_detail", ride_id=ride_id))
    finally:
        db.close()